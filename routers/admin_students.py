# routers/admin_students.py

from fastapi import APIRouter, Request, Depends, UploadFile, File, HTTPException, Form
from fastapi.responses import HTMLResponse, RedirectResponse, Response

import io
import openpyxl

from auth import load_user_dep
from context import ctx
from utils.permissions import has_permission
from utils.enums import PERM_GESTION_ALUMNOS

from db.students import (
    get_all_students,
    get_all_groups,
    create_student_if_not_exists,
)

router = APIRouter()


# ------------------------------------------------------
# Permiso
# ------------------------------------------------------

def _require_perm(user: dict):
    if not has_permission(user, PERM_GESTION_ALUMNOS):
        raise HTTPException(status_code=403)


# ------------------------------------------------------
# Vista principal
# ------------------------------------------------------

@router.get("/admin/students", response_class=HTMLResponse)
def admin_students(
    request: Request,
    user: dict = Depends(load_user_dep),
    grupo: str | None = None,
):
    _require_perm(user)

    groups = get_all_groups()

    if grupo:
        students = [
            s for s in get_all_students()
            if s["grupo"] == grupo
        ]
    else:
        students = get_all_students()

    return request.app.state.templates.TemplateResponse(
        "admin/students.html",
        ctx(
            request,
            user=user,
            title="Gestión de alumnos",
            students=students,
            groups=groups,
            selected_group=grupo,
        ),
    )


# ------------------------------------------------------
# Añadir alumno (formulario)
# ------------------------------------------------------

@router.post("/admin/students/create")
def create_student_post(
    request: Request,
    user: dict = Depends(load_user_dep),
    grupo: str = Form(...),
    alumno: str = Form(...),
):
    _require_perm(user)

    created = create_student_if_not_exists(
        grupo=grupo,
        alumno=alumno,
    )

    if not created:
        return RedirectResponse(
            "/admin/students?status=exists",
            status_code=303,
        )

    return RedirectResponse(
        "/admin/students?status=created",
        status_code=303,
    )


# ------------------------------------------------------
# Exportar Excel
# ------------------------------------------------------

@router.get("/admin/students/export")
def export_students(
    user: dict = Depends(load_user_dep),
):
    _require_perm(user)

    students = get_all_students()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Alumnos"

    ws.append(["Grupo", "Alumno"])

    for s in students:
        ws.append([s["grupo"], s["alumno"]])

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return Response(
        stream.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=alumnos.xlsx"
        },
    )


# ------------------------------------------------------
# Importar Excel
# ------------------------------------------------------

@router.post("/admin/students/import")
def import_students(
    user: dict = Depends(load_user_dep),
    file: UploadFile = File(...),
):
    _require_perm(user)

    if not file.filename.lower().endswith(".xlsx"):
        return RedirectResponse("/admin/students?status=error", status_code=303)

    try:
        wb = openpyxl.load_workbook(file.file)
        ws = wb.active
    except Exception:
        return RedirectResponse("/admin/students?status=error", status_code=303)

    headers = [cell.value for cell in ws[1]]
    expected = ["Grupo", "Alumno"]

    if headers[:2] != expected:
        return RedirectResponse("/admin/students?status=error", status_code=303)

    created = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        grupo, alumno = row[:2]

        if not grupo or not alumno:
            continue

        if create_student_if_not_exists(grupo=grupo, alumno=alumno):
            created += 1

    return RedirectResponse(
        f"/admin/students?status=imported&created={created}",
        status_code=303,
    )
