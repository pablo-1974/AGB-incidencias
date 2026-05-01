# routers/backup.py

from fastapi import (
    APIRouter, Request, Depends, HTTPException,
    UploadFile, File
)
from fastapi.responses import Response, HTMLResponse
from datetime import datetime
import io
import openpyxl

from auth import load_user_dep
from utils.permissions import has_permission
from utils.enums import PERM_BACKUP
from db.connection import get_db
from context import ctx

router = APIRouter()

# ======================================================
# PÁGINA BACKUP (UI)
# ======================================================
@router.get("/admin/backup", response_class=HTMLResponse)
def backup_page(
    request: Request,
    user: dict = Depends(load_user_dep),
):
    if not has_permission(user, PERM_BACKUP):
        raise HTTPException(status_code=403)

    return request.app.state.templates.TemplateResponse(
        "admin/backup.html",
        ctx(request, user=user, title="Copia de seguridad"),
    )

# ======================================================
# DESCARGA BACKUP
# ======================================================
@router.get("/admin/backup/download")
def backup_download(
    user: dict = Depends(load_user_dep),
):
    if not has_permission(user, PERM_BACKUP):
        raise HTTPException(status_code=403)

    wb = openpyxl.Workbook()

    # INFO
    ws_info = wb.active
    ws_info.title = "INFO"
    ws_info.append(["Aplicación", "Incidencias"])
    ws_info.append(["Fecha backup", datetime.now().strftime("%Y-%m-%d %H:%M")])
    ws_info.append(["Generado por", user["email"]])
    ws_info.append([])
    ws_info.append(["Contenido", "Copia completa de la base de datos"])

    # Exportar tablas
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT table_name
                FROM information_schema.tables
                WHERE table_schema = 'public'
                ORDER BY table_name
            """)
            tables = [r["table_name"] for r in cur.fetchall()]

            for table in tables:
                ws = wb.create_sheet(title=table)

                cur.execute(f'SELECT * FROM "{table}" LIMIT 1')
                columns = [d[0] for d in cur.description]
                ws.append(columns)

                cur.execute(f'SELECT * FROM "{table}"')
                for row in cur.fetchall():
                    ws.append(list(row.values()))

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = f"incidencias_backup_{datetime.now():%Y%m%d_%H%M}.xlsx"

    return Response(
        stream.read(),
        media_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )

# ======================================================
# SUBIDA BACKUP (IMPORTACIÓN INCREMENTAL)
# ======================================================
@router.post("/admin/backup/upload")
def backup_upload(
    file: UploadFile = File(...),
    user: dict = Depends(load_user_dep),
):
    # ---------------------------------
    # Seguridad
    # ---------------------------------
    if not has_permission(user, PERM_BACKUP):
        raise HTTPException(status_code=403)

    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Formato no válido")

    # ---------------------------------
    # Cargar Excel
    # ---------------------------------
    wb = openpyxl.load_workbook(file.file)

    # ---------------------------------
    # Importación incremental
    # ---------------------------------
    with get_db() as conn:
        with conn.cursor() as cur:

            # =================================
            # USERS
            # Duplicado: email
            # =================================
            if "users" in wb.sheetnames:
                ws = wb["users"]
                headers = [c.value for c in ws[1]]

                for row in ws.iter_rows(min_row=2, values_only=True):
                    data = dict(zip(headers, row))

                    # Ignorar usuarios ya existentes
                    cur.execute(
                        "SELECT 1 FROM users WHERE email = %s",
                        (data["email"],),
                    )
                    if cur.fetchone():
                        continue

                    # Insertar nuevo usuario
                    cur.execute(
                        """
                        INSERT INTO users (name, email, role)
                        VALUES (%s, %s, %s)
                        """,
                        (
                            data["name"],
                            data["email"],
                            data["role"],
                        ),
                    )

            # =================================
            # INCIDENTS
            # =================================
            if "incidents" in wb.sheetnames:
                ws = wb["incidents"]
                headers = [c.value for c in ws[1]]
            
                for row in ws.iter_rows(min_row=2, values_only=True):
                    data = dict(zip(headers, row))
            
                    # ---------------------------------
                    # Validación mínima (campos obligatorios)
                    # ---------------------------------
                    if not all([
                        data.get("teacher_id"),
                        data.get("fecha"),
                        data.get("hora"),
                        data.get("grupo"),
                        data.get("alumno"),
                        data.get("descripcion"),
                    ]):
                        continue  # ❌ fila incompleta → se ignora
            
                    # Comprobar si la incidencia ya existe
                    cur.execute(
                        """
                        SELECT 1
                        FROM incidents
                        WHERE teacher_id = %s
                          AND fecha = %s
                          AND hora = %s
                          AND grupo = %s
                          AND alumno = %s
                          AND descripcion = %s
                        """,
                        (
                            data["teacher_id"],
                            data["fecha"],
                            data["hora"],
                            data["grupo"],
                            data["alumno"],
                            data["descripcion"],
                        ),
                    )
                    if cur.fetchone():
                        continue
            
                    # Insertar nueva incidencia
                    cur.execute(
                        """
                        INSERT INTO incidents (
                            teacher_id,
                            teacher_name,
                            grupo,
                            alumno,
                            fecha,
                            hora,
                            hora_orden,
                            descripcion,
                            gravedad_inicial,
                            estado
                        )
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """,
                        (
                            data["teacher_id"],
                            data["teacher_name"],
                            data["grupo"],
                            data["alumno"],
                            data["fecha"],
                            data["hora"],
                            data["hora_orden"],
                            data["descripcion"],
                            data["gravedad_inicial"],
                            data["estado"],
                        ),
                    )

        # ✅ CONFIRMAR CAMBIOS
        conn.commit()

    return {
        "status": "ok",
        "message": "Importación completada correctamente"
    }
