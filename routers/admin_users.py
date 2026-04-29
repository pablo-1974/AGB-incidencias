# routers/admin_users.py
"""
Gestión de usuarios (ADMIN).

Funcionalidades:
- Listar usuarios
- Crear usuario (sin contraseña → primer login)
- Editar nombre, email y rol
- Activar / desactivar usuario
- Resetear contraseña (forzar primer login)

Acceso exclusivo para el rol admin.
Incluye salvaguardas para evitar dejar el sistema sin administradores.
"""

from fastapi import (
    APIRouter,
    Request,
    Form,
    HTTPException,
    UploadFile,
    File,
    Depends,
)
from fastapi.responses import HTMLResponse, RedirectResponse, Response

import io
import openpyxl

from auth import load_user_dep
from context import ctx
from utils.permissions import has_permission
from utils.enums import PERM_GESTION_USUARIOS, ROLES_TODOS

from db.users import (
    get_all_users,
    get_user_by_id,
    get_user_by_email,
    create_user_admin,
    update_user_admin,
    set_user_active,
    reset_user_password,
)

router = APIRouter()


# ----------------------------------------------------------------------
# UTILIDADES
# ----------------------------------------------------------------------

def _count_active_admins() -> int:
    """
    Devuelve el número de administradores activos.
    """
    from db.connection import get_db

    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT COUNT(*)
                FROM users
                WHERE role = 'admin'
                  AND active = 1
                """
            )
            return cur.fetchone()[0]


def _require_perm(user: dict):
    if not has_permission(user, PERM_GESTION_USUARIOS):
        raise HTTPException(status_code=403)


# ----------------------------------------------------------------------
# LISTADO DE USUARIOS
# ----------------------------------------------------------------------

@router.get("/admin/users", response_class=HTMLResponse)
def admin_users(
    request: Request,
    user: dict = Depends(load_user_dep),
):
    _require_perm(user)

    users = get_all_users()

    return request.app.state.templates.TemplateResponse(
        "admin/users.html",
        ctx(
            request,
            user=user,
            title="Gestión de usuarios",
            users=users,
            roles=sorted(ROLES_TODOS),
        ),
    )


# ----------------------------------------------------------------------
# CREAR USUARIO
# ----------------------------------------------------------------------

@router.post("/admin/users/create")
def admin_users_create(
    request: Request,
    user: dict = Depends(load_user_dep),
    name: str = Form(...),
    email: str = Form(...),
    role: str = Form(...),
):
    _require_perm(user)

    if role not in ROLES_TODOS:
        return RedirectResponse("/admin/users?status=error", status_code=303)

    create_user_admin(
        name=name.strip(),
        email=email.strip(),
        role=role,
        created_by=user["id"],
    )

    return RedirectResponse("/admin/users?status=created", status_code=303)


# ----------------------------------------------------------------------
# EDITAR USUARIO
# ----------------------------------------------------------------------

@router.post("/admin/users/update/{user_id}")
def admin_users_update(
    request: Request,
    user_id: int,
    user: dict = Depends(load_user_dep),
    name: str = Form(...),
    email: str = Form(...),
    role: str = Form(...),
):
    _require_perm(user)

    if role not in ROLES_TODOS:
        return RedirectResponse("/admin/users?status=error", status_code=303)

    target = get_user_by_id(user_id)
    if not target:
        return RedirectResponse("/admin/users?status=error", status_code=303)

    # Evitar quitar el último admin
    if target["role"] == "admin" and role != "admin":
        if _count_active_admins() <= 1:
            return RedirectResponse("/admin/users?status=error", status_code=303)

    update_user_admin(
        user_id=user_id,
        name=name.strip(),
        email=email.strip(),
        role=role,
    )

    return RedirectResponse("/admin/users?status=updated", status_code=303)


# ----------------------------------------------------------------------
# ACTIVAR / DESACTIVAR USUARIO
# ----------------------------------------------------------------------

@router.post("/admin/users/toggle/{user_id}")
def admin_users_toggle(
    request: Request,
    user_id: int,
    user: dict = Depends(load_user_dep),
):
    _require_perm(user)

    target = get_user_by_id(user_id)
    if not target:
        return RedirectResponse("/admin/users?status=error", status_code=303)

    # Evitar desactivar el último admin activo
    if target["role"] == "admin" and target["active"] == 1:
        if _count_active_admins() <= 1:
            return RedirectResponse("/admin/users?status=error", status_code=303)

    set_user_active(
        user_id=user_id,
        active=not bool(target["active"]),
    )

    return RedirectResponse("/admin/users?status=toggled", status_code=303)


# ----------------------------------------------------------------------
# RESET DE CONTRASEÑA
# ----------------------------------------------------------------------

@router.post("/admin/users/reset-password/{user_id}")
def admin_users_reset_password(
    request: Request,
    user_id: int,
    user: dict = Depends(load_user_dep),
):
    _require_perm(user)

    target = get_user_by_id(user_id)
    if not target:
        return RedirectResponse("/admin/users?status=error", status_code=303)

    reset_user_password(user_id=user_id)

    return RedirectResponse("/admin/users?status=reset", status_code=303)


# ----------------------------------------------------------------------
# IMPORTAR USUARIOS (EXCEL)
# ----------------------------------------------------------------------

@router.post("/admin/users/import")
def admin_users_import(
    user: dict = Depends(load_user_dep),
    file: UploadFile = File(...),
):
    _require_perm(user)

    if not file.filename.lower().endswith(".xlsx"):
        return RedirectResponse("/admin/users?status=error", status_code=303)

    try:
        wb = openpyxl.load_workbook(file.file)
        ws = wb.active
    except Exception:
        return RedirectResponse("/admin/users?status=error", status_code=303)

    headers = [cell.value for cell in ws[1]]
    expected = ["Nombre", "Email", "Rol"]

    if headers[:3] != expected:
        return RedirectResponse("/admin/users?status=error", status_code=303)

    created = 0
    updated = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        name, email, role = row[:3]

        if not email or not role or role not in ROLES_TODOS:
            continue

        email = email.strip()
        name = name.strip() if name else ""

        existing = get_user_by_email(email)

        if existing:
            update_user_admin(
                user_id=existing["id"],
                name=name or existing["name"],
                email=email,
                role=role,
            )
            updated += 1
        else:
            create_user_admin(
                name=name,
                email=email,
                role=role,
                created_by=user["id"],
            )
            created += 1

    return RedirectResponse(
        f"/admin/users?status=imported&created={created}&updated={updated}",
        status_code=303,
    )


# ----------------------------------------------------------------------
# EXPORTAR USUARIOS (EXCEL)
# ----------------------------------------------------------------------

@router.get("/admin/users/export")
def export_users(
    user: dict = Depends(load_user_dep),
):
    _require_perm(user)

    users = get_all_users()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Usuarios"

    ws.append([
        "Nombre",
        "Email",
        "Rol",
        "Activo",
        "Primer login pendiente",
        "Último acceso",
    ])

    for u in users:
        ws.append([
            u["name"],
            u["email"],
            u["role"],
            "Sí" if u["active"] == 1 else "No",
            "Sí" if u["must_change_password"] else "No",
            u["last_login_at"].strftime("%d/%m/%Y %H:%M")
                if u["last_login_at"] else "",
        ])

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return Response(
        stream.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=usuarios.xlsx"
        },
    )
